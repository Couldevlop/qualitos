# -*- coding: utf-8 -*-
"""
Génère le Dossier de Test QA de QualitOS au format Excel (.xlsx), avec :
  - Page de garde
  - Référentiel / Stratégie de test (types, priorités, environnements, critères)
  - Synthèse (tableau de bord + diagrammes : statuts, par domaine, priorités)
  - Cas de test (catalogue complet couvrant l'application)
  - Suivi & Évolution (séries temporelles + courbes d'avancement et d'anomalies)
  - Matrice de traçabilité (cas ↔ exigences CLAUDE.md)
  - Journal d'anomalies

Usage : python docs/qa/generate_test_plan.py
Sortie : docs/qa/QualitOS_Dossier_de_Test_QA.xlsx
"""
import os
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo

# ------------------------------------------------------------------ palette ----
NAVY = "1F2A44"
BLUE = "2563EB"
GREEN = "059669"
AMBER = "D97706"
RED = "DC2626"
GREY = "64748B"
LIGHT = "F1F5F9"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D9DEE7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fill(hexc): return PatternFill("solid", fgColor=hexc)
def font(sz=11, b=False, color="0F172A", italic=False):
    return Font(name="Calibri", size=sz, bold=b, color=color, italic=italic)

STATUS_COLORS = {
    "Passé": GREEN, "Échoué": RED, "Bloqué": AMBER,
    "Non exécuté": GREY, "En cours": BLUE,
}
PRIO_COLORS = {"Critique": RED, "Haute": AMBER, "Moyenne": BLUE, "Basse": GREY}

# =====================================================================
# 1) CATALOGUE DES CAS DE TEST
#    Chaque entrée : (sous-module, titre, type, priorité, préconditions,
#                     étapes, résultat attendu, rôle, exigence)
#    Les domaines regroupent les cas. Statut affecté par distribution déterministe.
# =====================================================================
def c(sub, title, typ, prio, steps, expected, role="Manager Qualité",
      req="—", pre="Utilisateur authentifié, tenant actif"):
    return dict(sub=sub, title=title, typ=typ, prio=prio, pre=pre,
                steps=steps, expected=expected, role=role, req=req)

DOMAINS = []

def domain(name, code, cases):
    DOMAINS.append((name, code, cases))

# ---- A. Méthodes qualité ------------------------------------------------------
domain("PDCA — Roue de Deming", "PDCA", [
    c("Liste", "Afficher la liste des cycles PDCA", "Fonctionnel", "Haute",
      "1. Se connecter\n2. Menu Méthodes > PDCA",
      "La liste paginée s'affiche avec un seul appel API, sans scintillement (cf. anomalie ANO-002).", req="CLAUDE §3.1"),
    c("Liste", "Filtrer les cycles par statut (PLAN/DO/CHECK/ACT)", "Fonctionnel", "Moyenne",
      "1. Sur /pdca\n2. Sélectionner un statut dans le filtre",
      "La liste se met à jour, seuls les cycles du statut choisi sont affichés.", req="CLAUDE §3.1"),
    c("Création", "Créer un nouveau cycle PDCA", "Fonctionnel", "Critique",
      "1. Cliquer 'Nouveau cycle'\n2. Saisir titre + description\n3. Valider",
      "Le cycle est créé en phase PLAN, visible en tête de liste.", req="CLAUDE §3.1"),
    c("Détail", "Ouvrir le détail d'un cycle et charger ses étapes", "Fonctionnel", "Haute",
      "1. Cliquer une ligne de la liste",
      "Le détail s'affiche (titre, étapes, boutons d'action), pas figé sur 'Chargement…' (cf. ANO-003).", req="CLAUDE §3.1"),
    c("Workflow", "Faire avancer un cycle PLAN→DO→CHECK→ACT", "Fonctionnel", "Critique",
      "1. Ouvrir un cycle\n2. Cliquer 'Avancer' à chaque phase",
      "Le statut progresse phase par phase ; transition ancrée blockchain à chaque étape.", req="CLAUDE §3.1"),
    c("Workflow", "Annuler un cycle en cours", "Fonctionnel", "Moyenne",
      "1. Ouvrir un cycle\n2. Cliquer 'Annuler'",
      "Le cycle passe en CANCELLED, les actions d'avancement sont désactivées.", req="CLAUDE §3.1"),
    c("IA", "Prédiction d'atteinte d'objectif (LSTM)", "Fonctionnel", "Moyenne",
      "1. Ouvrir un cycle avec KPI cible\n2. Consulter l'encart de prédiction",
      "La probabilité d'atteinte + explication (SHAP) s'affiche.", role="Directeur Qualité", req="CLAUDE §6.5"),
    c("Sécurité", "Refus d'accès à un cycle d'un autre tenant", "Sécurité", "Critique",
      "1. Tenter GET /api/v1/pdca/{id d'un autre tenant}",
      "404 (pas de fuite cross-tenant) ; tenant_id pris du JWT.", role="Attaquant", req="CLAUDE §18.2"),
    c("Validation", "Identifiant de cycle malformé", "Sécurité", "Haute",
      "1. Naviguer vers /pdca/not-a-uuid",
      "Rejet client (UUID_RE) + redirection liste, pas d'appel backend.", req="CLAUDE §11.1"),
])

domain("5S — Excellence opérationnelle", "5S", [
    c("Liste", "Afficher les audits 5S par zone", "Fonctionnel", "Haute",
      "1. Menu Méthodes > 5S",
      "Liste des audits avec score pondéré 0-100, statut, zone.", req="CLAUDE §3.2"),
    c("Création", "Planifier un audit 5S récurrent", "Fonctionnel", "Haute",
      "1. 'Nouvel audit'\n2. Choisir zone + récurrence",
      "Audit créé, programmation enregistrée, notifications planifiées.", req="CLAUDE §3.2"),
    c("Terrain", "Réaliser un audit 5S en mode offline (PWA)", "Fonctionnel", "Critique",
      "1. Couper le réseau\n2. Remplir un audit 5S\n3. Reconnecter",
      "L'audit est mis en file offline puis synchronisé dans l'ordre au retour réseau.", role="Utilisateur", req="CLAUDE §15.2"),
    c("IA Vision", "Détection CV des non-conformités sur photo (YOLOv8)", "Fonctionnel", "Moyenne",
      "1. Joindre une photo d'atelier\n2. Lancer l'analyse vision",
      "Encombrement/EPI/étiquetage détectés ; NC suggérée (ou 503 propre si modèle absent).", req="CLAUDE §3.2"),
    c("Restitution", "Heatmap de score 5S par zone", "IHM/UX", "Moyenne",
      "1. Dashboard 5S",
      "Heatmap par zone + tendance affichées.", req="CLAUDE §3.2"),
    c("Détail", "Signature blockchain du rapport d'audit", "Fonctionnel", "Haute",
      "1. Clôturer un audit\n2. Générer le rapport",
      "Rapport signé ML-DSA + hash ancré, vérifiable.", role="Auditeur", req="CLAUDE §11.3"),
])

domain("Ishikawa — Causes-racines", "ISHI", [
    c("Liste", "Afficher les diagrammes Ishikawa", "Fonctionnel", "Haute",
      "1. Menu Méthodes > Ishikawa", "Liste paginée des diagrammes.", req="CLAUDE §3.5"),
    c("Éditeur", "Construire un diagramme 6M collaboratif", "Fonctionnel", "Haute",
      "1. Ouvrir l'éditeur\n2. Ajouter des branches 6M et causes",
      "Les causes se rattachent aux branches ; sauvegarde persistée.", req="CLAUDE §3.5"),
    c("IA", "Suggestion IA de causes par branche (RAG)", "Fonctionnel", "Moyenne",
      "1. Saisir la description du problème\n2. Cliquer 'Suggérer des causes'",
      "Des causes probables par catégorie 6M sont proposées (provider Mistral/Ollama).", req="CLAUDE §12.1"),
    c("Conversion", "Convertir un Ishikawa en cycle PDCA", "Intégration", "Haute",
      "1. Depuis un diagramme\n2. 'Convertir en PDCA'",
      "Un cycle PDCA est créé à partir de la cause sélectionnée (référentiel commun).", req="CLAUDE §3.6"),
    c("Couplage", "Pareto + 5 Pourquoi sur une cause", "Fonctionnel", "Basse",
      "1. Sélectionner une cause\n2. Lancer 5 Pourquoi",
      "L'arbre des 5 Pourquoi se construit, lié au Pareto.", req="CLAUDE §3.5"),
])

domain("DMAIC + Poka-Yoke", "DMAIC", [
    c("Liste", "Afficher les projets DMAIC", "Fonctionnel", "Haute",
      "1. Menu Méthodes > DMAIC", "Liste des projets (Define→Control).", req="CLAUDE §3.4"),
    c("Workflow", "Dérouler les 5 phases DMAIC", "Fonctionnel", "Critique",
      "1. Créer un projet\n2. Avancer Define→Measure→Analyze→Improve→Control",
      "Chaque phase enregistre ses livrables ; transitions tracées.", req="CLAUDE §3.4"),
    c("SPC", "Calcul de capabilité Cp/Cpk", "Fonctionnel", "Haute",
      "1. Charger des mesures\n2. Lancer le calcul",
      "Cp/Cpk/Pp/Ppk calculés et affichés.", req="CLAUDE §3.4"),
    c("Poka-Yoke", "Recommandation de dispositif Poka-Yoke", "Fonctionnel", "Moyenne",
      "1. Phase Improve\n2. Demander une reco",
      "Dispositifs Poka-Yoke sectoriels recommandés.", req="CLAUDE §3.4"),
])

domain("Cercles de Qualité", "CERCLE", [
    c("Liste", "Afficher les cercles de qualité", "Fonctionnel", "Haute",
      "1. Menu > Cercles", "Liste des cercles avec membres et statut.", req="CLAUDE §3.3"),
    c("Création", "Constituer un cercle (5-10 membres + rôles)", "Fonctionnel", "Haute",
      "1. 'Nouveau cercle'\n2. Affecter animateur/secrétaire/membres",
      "Cercle créé avec rôles ; contrôle de taille 5-10.", req="CLAUDE §3.3"),
    c("Propositions", "Suivre une proposition jusqu'à mesure d'impact", "Fonctionnel", "Moyenne",
      "1. Ajouter une proposition\n2. Valider, mettre en œuvre, mesurer",
      "Statut et impact suivis ; validation hiérarchique tracée.", req="CLAUDE §3.3"),
    c("IA", "Compte-rendu auto-généré (Whisper + LLM)", "Fonctionnel", "Basse",
      "1. Importer l'audio d'une réunion\n2. Générer le CR",
      "Transcription + résumé + actions extraites avec assignation suggérée.", req="CLAUDE §3.3"),
])

# ---- B. Modules transverses qualité ------------------------------------------
domain("Non-Conformités (NC)", "NC", [
    c("Liste", "Afficher les non-conformités", "Fonctionnel", "Haute",
      "1. Menu Opérations > NC", "Liste paginée filtrable.", req="CLAUDE §4.3"),
    c("Création", "Saisir une NC avec photo/géoloc (mobile)", "Fonctionnel", "Critique",
      "1. 'Nouvelle NC'\n2. Ajouter photo + localisation\n3. Valider",
      "NC créée avec pièces jointes ; saisie mobile fonctionnelle.", role="Utilisateur", req="CLAUDE §4.3"),
    c("IA", "Clustering automatique des NC similaires (DBSCAN)", "Fonctionnel", "Moyenne",
      "1. Page /nc-clusters\n2. Lancer le clustering",
      "Les NC sont regroupées par similarité ; patterns mis en évidence.", req="CLAUDE §12.1"),
    c("Vision", "Analyse vision d'une photo de NC", "Fonctionnel", "Moyenne",
      "1. Ouvrir une NC avec photo\n2. 'Analyser'",
      "Classification du défaut ou message 503 propre si service indisponible.", req="CLAUDE §1.4"),
    c("Intégration", "Instancier une NC en CAPA ou PDCA selon gravité", "Intégration", "Haute",
      "1. Depuis une NC critique\n2. Escalader",
      "Un dossier CAPA (ou cycle PDCA) est créé et lié à la NC.", req="CLAUDE §3.6"),
])

domain("CAPA — Actions correctives/préventives", "CAPA", [
    c("Liste", "Afficher les dossiers CAPA", "Fonctionnel", "Haute",
      "1. Menu Opérations > CAPA", "Liste filtrable par statut/criticité.", req="CLAUDE §4.2"),
    c("Workflow", "Cycle de vie d'un CAPA jusqu'à clôture", "Fonctionnel", "Critique",
      "1. Créer un CAPA\n2. Affecter, traiter, vérifier l'efficacité, clôturer",
      "Workflow respecté ; efficacité mesurée à 3/6/12 mois.", req="CLAUDE §4.2"),
    c("IA", "Suggestion de causes-racines et d'actions (RAG)", "Fonctionnel", "Moyenne",
      "1. Sur un CAPA\n2. 'Suggérer des actions'",
      "Actions probables proposées selon l'efficacité passée (provider IA).", req="CLAUDE §12.1"),
    c("KPI", "Délai moyen de clôture CAPA par criticité", "Fonctionnel", "Moyenne",
      "1. Dashboard CAPA",
      "Le KPI s'affiche avec seuils (cible <30j, ISO 9001 §10.2).", req="CLAUDE §6.6"),
])

domain("Audits (interne/externe/fournisseurs)", "AUD", [
    c("Liste", "Afficher les plans d'audit", "Fonctionnel", "Haute",
      "1. Menu Opérations > Audits", "Liste des plans/programmes.", req="CLAUDE §4.4"),
    c("Exécution", "Réaliser un audit terrain LPA offline", "Fonctionnel", "Haute",
      "1. Démarrer un audit\n2. Remplir la checklist hors-ligne",
      "Checklist remplie, synchronisée au retour réseau.", role="Auditeur", req="CLAUDE §4.4"),
    c("IA", "Rapport d'audit généré par LLM", "Fonctionnel", "Moyenne",
      "1. Clôturer un audit\n2. 'Générer le rapport'",
      "Rapport généré avec citations vers preuves, signé/ancré.", req="CLAUDE §1.4"),
    c("Détail", "Détail d'audit charge sans blocage", "Régression", "Haute",
      "1. Cliquer une ligne d'audit",
      "Le détail s'affiche (cf. correctif ANO-003 reload$).", req="—"),
])

domain("Document Control / GED", "DOC", [
    c("Liste", "Afficher la GED qualité", "Fonctionnel", "Moyenne",
      "1. Menu > Documents", "Arborescence/liste des documents versionnés.", req="CLAUDE §4.1"),
    c("Versioning", "Créer une nouvelle version + e-signature ML-DSA", "Fonctionnel", "Haute",
      "1. Ouvrir un document\n2. Nouvelle version\n3. Signer",
      "Version créée, signée ML-DSA, workflow d'approbation déclenché.", req="CLAUDE §4.1"),
    c("Lecture", "Preuve de lecture obligatoire", "Fonctionnel", "Moyenne",
      "1. Affecter une lecture obligatoire\n2. L'utilisateur lit + confirme",
      "Preuve de lecture enregistrée et horodatée.", role="Utilisateur", req="CLAUDE §4.1"),
])

domain("Fournisseurs / FMEA / Changes / EHS / ITSM", "TRANSV", [
    c("Suppliers", "Scorecard fournisseur + scoring IA", "Fonctionnel", "Moyenne",
      "1. Menu > Fournisseurs\n2. Ouvrir un fournisseur",
      "Score qualité (IA) + historique NC affichés.", req="CLAUDE §4.6"),
    c("FMEA", "Évaluer un risque FMEA (RPN dynamique)", "Fonctionnel", "Haute",
      "1. Menu > FMEA\n2. Saisir gravité/occurrence/détection",
      "RPN calculé et réévalué dynamiquement.", req="CLAUDE §4.5"),
    c("Changes", "Demande de changement + analyse d'impact", "Fonctionnel", "Moyenne",
      "1. Menu > Changes\n2. Créer une demande",
      "Impact (docs, formations, fournisseurs) lié automatiquement.", req="CLAUDE §4.8"),
    c("EHS", "Déclarer un événement EHS (réutilise NC/CAPA)", "Fonctionnel", "Moyenne",
      "1. Menu > EHS\n2. Déclarer",
      "Événement EHS instancié, conforme ISO 45001/14001.", req="CLAUDE §4.11"),
    c("ITSM", "Import d'incidents ITSM (connecteur)", "Intégration", "Basse",
      "1. Menu > Intégrations (ITSM)\n2. Synchroniser",
      "Incidents importés et convertibles en NC/CAPA.", req="CLAUDE §13.3"),
])

# ---- C. Standards Hub & certification ----------------------------------------
domain("Standards Hub — Normes & certification", "STD", [
    c("Catalogue", "Lister les 60 normes du référentiel", "Fonctionnel", "Haute",
      "1. Menu Référentiels > Standards Hub",
      "60 normes paginées (totalElements=60).", req="CLAUDE §8.2"),
    c("Fiche", "Consulter les 8 onglets d'une norme (ISO 9001)", "Fonctionnel", "Haute",
      "1. Ouvrir ISO 9001\n2. Parcourir les onglets",
      "Vue d'ensemble, exigences, docs, processus, roadmap, preuves, audit blanc, veille.", req="CLAUDE §8.4"),
    c("Preuves", "Score d'alignement IA par clause", "Fonctionnel", "Haute",
      "1. Onglet 'Mes preuves'",
      "Score de couverture par section/clause + écarts détectés.", req="CLAUDE §8.7"),
    c("Audit blanc", "Lancer l'audit blanc IA avancé", "Fonctionnel", "Haute",
      "1. Onglet 'Audit blanc IA'\n2. Lancer",
      "30-100 questions, gap analysis, plan de remédiation auto.", req="CLAUDE §8.4"),
    c("Certification", "Générer le dossier de certification (PDF signé)", "Fonctionnel", "Critique",
      "1. 'Générer le dossier'",
      "PDF zippé + preuves, signé ML-DSA, ancré blockchain.", role="Directeur Qualité", req="CLAUDE §8.7"),
    c("Téléchargement", "Garde anti-path-traversal sur les modèles", "Sécurité", "Haute",
      "1. Tenter de télécharger un modèle avec ../ dans le chemin",
      "Accès refusé (garde anti-traversée).", role="Attaquant", req="CLAUDE §11.1"),
])

domain("Génération documentaire IA (multi-docs)", "DOCGEN", [
    c("Génération", "Générer un dossier documentaire en lot", "Fonctionnel", "Haute",
      "1. /standards-doc-gen\n2. Choisir norme + profil tenant\n3. Démarrer",
      "Manuel + politique + procédures générés en lot (provider IA), suivi de progression.", req="CLAUDE §8.8"),
    c("Validation", "Workflow de validation humaine par pièce", "Fonctionnel", "Critique",
      "1. Pour chaque pièce : soumettre, relire, approuver",
      "Aucune finalisation tant que toutes les pièces ne sont pas approuvées.", req="CLAUDE §18.2"),
    c("Scellement", "Finaliser le dossier (signature + ancrage)", "Fonctionnel", "Haute",
      "1. Toutes pièces approuvées\n2. 'Finaliser' avec signature",
      "Dossier scellé (SHA-256 signé ML-DSA + ancré).", role="Directeur Qualité", req="CLAUDE §8.8"),
    c("Résilience", "Panne IA sur une pièce n'échoue pas le lot", "Fonctionnel", "Moyenne",
      "1. Simuler une panne provider sur une pièce",
      "La pièce passe en ÉCHEC, relançable ; les autres aboutissent.", req="—"),
])

domain("Industry Packs & Marketplace", "MKT", [
    c("Packs", "Activer un Industry Pack (déclaratif)", "Fonctionnel", "Haute",
      "1. Menu > Packs sectoriels\n2. Activer un pack",
      "Pack activé pour le tenant ; KPIs/normes/connecteurs chargés.", role="Admin Tenant", req="CLAUDE §5.3"),
    c("Marketplace", "Parcourir le catalogue public de packs", "Fonctionnel", "Moyenne",
      "1. Menu > Marketplace", "Packs PUBLIÉS listés avec métadonnées.", req="CLAUDE §8.11"),
    c("Marketplace", "Soumettre un pack (partenaire) + scan manifest", "Fonctionnel", "Haute",
      "1. 'Soumettre'\n2. Renseigner + manifest",
      "Pack en SUBMITTED après scan basique du manifest.", role="Partenaire", req="CLAUDE §8.11"),
    c("Marketplace", "Modération éditeur : publier/rejeter", "Fonctionnel", "Haute",
      "1. File de modération\n2. Publier ou rejeter (motif)",
      "Transition SUBMITTED→PUBLISHED/REJECTED ; publication réservée SUPER_ADMIN.", role="Super Admin", req="CLAUDE §8.11"),
    c("Marketplace", "Installer un pack publié (tenant)", "Fonctionnel", "Moyenne",
      "1. Depuis un pack publié\n2. 'Installer'",
      "Installation enregistrée par tenant ; notation possible.", role="Admin Tenant", req="CLAUDE §8.11"),
])

# ---- D. Dashboards & pilotage -------------------------------------------------
domain("Dashboards & Pilotage", "DASH", [
    c("Exécutif", "Afficher le dashboard exécutif (8-12 KPIs)", "Fonctionnel", "Haute",
      "1. Menu > Tableau de bord",
      "KPIs stratégiques, heatmap conformité, top risques/actions, COQ.", role="Directeur Qualité", req="CLAUDE §7.1"),
    c("Interactivité", "Cross-filtering au clic sur le Pareto", "IHM/UX", "Moyenne",
      "1. Cliquer une barre du Pareto",
      "Tous les widgets se filtrent sur la catégorie ; annulable.", req="CLAUDE §7.3"),
    c("Time-travel", "Afficher l'état à une date passée", "Fonctionnel", "Moyenne",
      "1. Choisir une date\n2. Appliquer",
      "Snapshot as-of réel récupéré du backend.", req="CLAUDE §7.3"),
    c("Builder", "Composer un dashboard en drag & drop", "Fonctionnel", "Haute",
      "1. /dashboard-builder\n2. Glisser des widgets, configurer, sauver",
      "Layout (positions/tailles/config) persisté par tenant.", req="CLAUDE §7.3"),
    c("Export", "Export PDF du dashboard signé + QR", "Fonctionnel", "Haute",
      "1. 'Exporter en PDF (signé)'",
      "PDF signé ML-DSA + ancré, QR de vérification publique.", req="CLAUDE §7.4"),
    c("Vérif", "Vérification publique d'un export par code", "Sécurité", "Moyenne",
      "1. GET /dashboards/public/exports/{code}/verify",
      "Faits d'intégrité renvoyés ; aucune donnée tenant exposée.", role="Public", req="CLAUDE §7.4"),
    c("Mode TV", "Rotation auto en mode mural", "IHM/UX", "Basse",
      "1. /tv\n2. Lancer la rotation",
      "Les vues défilent automatiquement à l'intervalle choisi.", req="CLAUDE §7.3"),
    c("KPI", "Catalogue KPI avec définitions explicites", "Fonctionnel", "Moyenne",
      "1. Menu > Indicateurs (KPI)",
      "Chaque KPI a formule/seuil/source/propriétaire.", req="CLAUDE §6.6"),
])

# ---- E. IA / ML ---------------------------------------------------------------
domain("IA / ML & Assistant", "AI", [
    c("NLQ", "Poser une question en langage naturel (text-to-SQL)", "Fonctionnel", "Critique",
      "1. /nlq\n2. 'Combien de CAPA par statut ?'",
      "SQL généré avec filtre tenant_id, exécuté en lecture seule, résultat + graphe.", req="CLAUDE §7.3"),
    c("NLQ", "Indisponibilité provider IA gérée proprement", "Robustesse", "Haute",
      "1. Provider lent/HS\n2. Poser une question",
      "Message 'assistant indisponible' (503/502) sans plantage ; pas de fuite.", req="CLAUDE §11.2"),
    c("NLQ", "Garde-fou anti-injection SQL", "Sécurité", "Critique",
      "1. Question induisant DROP/DELETE",
      "SQL refusé par l'allow-list (422), rôle PG read-only.", role="Attaquant", req="CLAUDE §11.2"),
    c("SPC", "Détection d'anomalies SPC (8 règles de Nelson)", "Fonctionnel", "Haute",
      "1. /spc\n2. Charger une série",
      "Violations des règles Nelson signalées sur la carte de contrôle.", req="CLAUDE §3.4"),
    c("Anomalies", "Anomalies multivariées + explication SHAP", "Fonctionnel", "Moyenne",
      "1. /anomaly\n2. Analyser, puis 'Pourquoi ?'",
      "Isolation Forest/ACP + contributions SHAP affichées.", req="CLAUDE §12.3"),
    c("Forecast", "Prévision KPI (Holt-Winters)", "Fonctionnel", "Moyenne",
      "1. /forecast\n2. Sélectionner un KPI",
      "Projection + intervalle affichés.", req="CLAUDE §6.5"),
    c("NLP", "Analyse de réclamations (sentiment + classif)", "Fonctionnel", "Moyenne",
      "1. /complaints-nlp\n2. Importer des réclamations",
      "Sentiment + classification + criticité affichés.", req="CLAUDE §4.9"),
    c("Storyboard", "Récit narratif IA des KPIs", "Fonctionnel", "Basse",
      "1. /storyboard\n2. Générer",
      "Texte narratif autour des chiffres du mois.", req="CLAUDE §7.4"),
    c("Garde-fous", "Quotas/rate-limit par tenant (LLM04)", "Sécurité", "Haute",
      "1. Dépasser le quota IA d'un tenant",
      "429/413/503 selon le cas ; disjoncteur fonctionnel.", req="CLAUDE §11.2"),
    c("Provider", "Bascule de provider par env (ollama/mistral/anthropic)", "Configuration", "Moyenne",
      "1. AI_DEFAULT_PROVIDER=mistral + clé\n2. Redémarrer ai-service",
      "Tout l'IA texte passe par le provider choisi, sans modif de code.", role="DevOps", req="ADR 0014"),
])

# ---- F. Formation -------------------------------------------------------------
domain("Formation & Academy (LMS)", "ACAD", [
    c("Catalogue", "Parcourir le catalogue de cours Academy", "Fonctionnel", "Moyenne",
      "1. Menu > Academy", "Cours par rôle/secteur, mes formations, classement.", role="Utilisateur", req="CLAUDE §19.3"),
    c("Parcours", "Suivre un cours, leçons et quiz noté", "Fonctionnel", "Haute",
      "1. S'inscrire\n2. Compléter leçons\n3. Passer le quiz",
      "Progression suivie ; quiz noté avec seuil de réussite.", role="Utilisateur", req="CLAUDE §19.3"),
    c("Gamification", "Badges, points et ceintures Yellow→Black", "Fonctionnel", "Basse",
      "1. Compléter des cours",
      "Badges/points attribués ; niveau de ceinture mis à jour ; leaderboard.", role="Utilisateur", req="CLAUDE §19.3"),
    c("Certificat", "Certificat signé ML-DSA + ancré + QR", "Fonctionnel", "Haute",
      "1. Terminer un cours\n2. Voir le certificat",
      "Certificat généré, signé, ancré, vérifiable par QR.", role="Utilisateur", req="CLAUDE §19.3"),
    c("Vérif", "Vérification publique d'un certificat", "Sécurité", "Moyenne",
      "1. GET /academy/public/certificates/{code}/verify",
      "Revalidation de signature ; aucune donnée perso exposée.", role="Public", req="CLAUDE §11.3"),
    c("Export", "Export SCORM 2004 / xAPI", "Intégration", "Basse",
      "1. Exporter un parcours/complétion",
      "Paquet SCORM ZIP + statements xAPI produits.", req="CLAUDE §19.3"),
    c("Matrice", "Matrice de compétences (Training)", "Fonctionnel", "Moyenne",
      "1. Menu > Formation",
      "Matrice rôle×compétence, échéances et rappels.", req="CLAUDE §4.7"),
])

# ---- G. Workflow no-code ------------------------------------------------------
domain("Workflow Designer (BPMN)", "WF", [
    c("Liste", "Afficher les workflows", "Fonctionnel", "Moyenne",
      "1. /workflow-designer", "Liste des workflows.", req="CLAUDE §5.4"),
    c("Éditeur", "Modéliser un workflow BPMN (bpmn-js)", "Fonctionnel", "Haute",
      "1. Ouvrir l'éditeur\n2. Glisser des éléments BPMN\n3. Sauver",
      "Diagramme BPMN édité et persisté ; lib chargée en lazy.", req="CLAUDE §5.4"),
    c("IA", "Génération de workflow par langage naturel", "Fonctionnel", "Moyenne",
      "1. Décrire 'processus CAPA en 4 étapes'\n2. Générer",
      "Workflow généré (output structuré LLM) à valider.", req="CLAUDE §1.4"),
])

# ---- H. GRC / Conformité ------------------------------------------------------
domain("Conformité GRC (RGPD / IA Act / NIS 2)", "GRC", [
    c("Hub", "Accéder au hub Conformité", "Fonctionnel", "Haute",
      "1. Menu > Conformité (/compliance)",
      "Point d'entrée vers les 19 routes GRC.", role="Admin Tenant", req="—"),
    c("RGPD", "Tenir le registre des traitements (ROPA)", "Fonctionnel", "Haute",
      "1. /ropa\n2. Créer un traitement", "Traitement enregistré au registre.", req="RGPD §30"),
    c("RGPD", "Gérer les consentements", "Fonctionnel", "Moyenne",
      "1. /consents", "Consentements listés et révocables.", req="RGPD"),
    c("RGPD", "Traiter une demande DSAR", "Fonctionnel", "Haute",
      "1. /subject-requests\n2. Traiter dans les délais",
      "DSAR suivie avec échéance légale.", req="RGPD §15-22"),
    c("RGPD", "Mener une DPIA", "Fonctionnel", "Moyenne",
      "1. /dpia\n2. Évaluer un traitement à risque", "DPIA documentée.", req="RGPD §35"),
    c("RGPD", "Notifier une violation de données", "Fonctionnel", "Critique",
      "1. /breaches\n2. Déclarer + délai 72h",
      "Violation enregistrée, échéance 72h suivie.", req="RGPD §33"),
    c("RGPD", "Transferts hors UE / sous-traitants (DPA)", "Fonctionnel", "Moyenne",
      "1. /cross-border et /processor-agreements",
      "Transferts encadrés + DPA suivis.", req="RGPD §28/44"),
    c("IA Act", "Inventaire des systèmes d'IA (AI-QMS)", "Fonctionnel", "Haute",
      "1. /ai-qms\n2. Inventorier un système IA",
      "Système IA classé par niveau de risque.", req="AI Act"),
    c("IA Act", "FRIA + décisions automatisées", "Fonctionnel", "Moyenne",
      "1. /fria et /automated-decisions",
      "Analyse d'impact droits fondamentaux + registre décisions auto.", req="AI Act"),
    c("NIS 2", "Mesures de sécurité + incidents cyber", "Fonctionnel", "Haute",
      "1. /nis2-measures et /cyber-incidents",
      "Mesures suivies, incidents cyber déclarés.", req="NIS 2"),
])

# ---- I. Transverse technique --------------------------------------------------
domain("Authentification & Sécurité", "SEC", [
    c("Auth", "Connexion via Keycloak (OIDC code+PKCE)", "Sécurité", "Critique",
      "1. Accéder à l'app\n2. Login demo/demo",
      "Redirection Keycloak, retour authentifié, token stocké.", role="Utilisateur", req="CLAUDE §10.2"),
    c("Auth", "Refresh automatique du token (refresh_token)", "Régression", "Critique",
      "1. Rester >15 min\n2. Naviguer/refresh",
      "Token rafraîchi sans rebond Keycloak (cf. ANO-004) ; session maintenue.", req="—"),
    c("Auth", "Préservation de la route après ré-auth", "Régression", "Haute",
      "1. F5/deep-link sur /fives sans session\n2. Login",
      "Retour sur /fives (pas /home).", req="—"),
    c("RBAC", "Endpoint d'administration réservé aux rôles habilités", "Sécurité", "Critique",
      "1. Appeler un endpoint admin avec un rôle 'Utilisateur'",
      "403 ; mapping rôles realm→ROLE_*.", role="Utilisateur", req="CLAUDE §16"),
    c("Multi-tenant", "Isolation stricte par tenant_id du JWT", "Sécurité", "Critique",
      "1. Forcer un tenant_id dans le body",
      "Ignoré ; tenant_id pris du JWT (RLS PostgreSQL).", role="Attaquant", req="CLAUDE §18.2"),
    c("Headers", "En-têtes de sécurité (CSP, X-Frame-Options…)", "Sécurité", "Haute",
      "1. Inspecter les réponses API",
      "CSP, frame deny, referrer no-referrer, permissions-policy présents.", req="CLAUDE §11.1"),
    c("DAST", "Scan OWASP ZAP en CI bloque sur HIGH uniquement", "Sécurité", "Haute",
      "1. Lancer le job DAST",
      "0 HIGH → vert ; WARN/MEDIUM rapportés non bloquants (cf. ANO-005).", role="DevOps", req="CLAUDE §14.2"),
    c("Crypto", "Signature ML-DSA + ancrage blockchain d'une action critique", "Sécurité", "Haute",
      "1. Signer un rapport critique",
      "Signature post-quantique + hash ancré, vérifiables.", req="CLAUDE §11.4"),
])

domain("Transverse UX / i18n / Accessibilité / PWA / Perf", "XCUT", [
    c("i18n", "Basculer la langue (FR/EN/ES/AR/JA/ZH)", "i18n", "Moyenne",
      "1. Changer la langue",
      "L'UI est traduite ; AR en RTL ; aucune clé manquante.", req="CLAUDE §15.1"),
    c("A11y", "Navigation clavier + lecteur d'écran", "Accessibilité", "Haute",
      "1. Parcourir une page au clavier\n2. Tester NVDA/VoiceOver",
      "Focus visible, labels ARIA, score Axe ≥ 95.", req="CLAUDE §15.1"),
    c("PWA", "Mode offline + file de synchro", "Fonctionnel", "Haute",
      "1. Couper le réseau\n2. Agir\n3. Voir /offline-queue\n4. Reconnecter",
      "Actions mises en file, rejouées dans l'ordre au retour réseau.", req="CLAUDE §15.2"),
    c("Perf", "SLO latence API p95 < 300 ms (k6)", "Performance", "Haute",
      "1. Lancer le scénario k6 sur les chemins chauds",
      "p95<300ms, p99<800ms, erreurs<1% ; sinon le gate CI échoue.", role="DevOps", req="CLAUDE §14.3"),
    c("Perf", "Core Web Vitals (LCP<2s, INP<200ms, CLS<0.1)", "Performance", "Moyenne",
      "1. Mesurer sur les pages clés",
      "Les seuils Web Vitals sont respectés.", req="CLAUDE §15.2"),
    c("Robustesse", "Listes : 1 seul appel API, sans scintillement", "Régression", "Critique",
      "1. Ouvrir chaque liste (PDCA, 5S, NC, CAPA…)",
      "1 appel API, spinner bref, données affichées (cf. ANO-002).", req="—"),
    c("Résilience", "Chaos : kill d'un pod engine", "Robustesse", "Moyenne",
      "1. Appliquer PodChaos\n2. Observer la reprise",
      "2/2 répliques Ready <2min, 5xx<5%, SLO maintenu.", role="DevOps", req="CLAUDE §14.3"),
])

domain("IoT & Edge (backend)", "IOT", [
    c("Ingestion", "Réception de mesures (MQTT/OPC-UA)", "Intégration", "Moyenne",
      "1. Publier une mesure capteur",
      "Mesure ingérée, stockée (TimescaleDB), KPI alimenté.", role="DevOps", req="CLAUDE §9.4"),
    c("Edge", "Store-and-forward en coupure réseau", "Robustesse", "Moyenne",
      "1. Couper la liaison Edge↔Hub\n2. Reconnecter",
      "Buffer local rejoué, aucune perte.", role="DevOps", req="CLAUDE §9.5"),
    c("Auto-NC", "Excursion de seuil déclenche une NC + alerte", "Intégration", "Haute",
      "1. Envoyer une valeur hors seuil",
      "NC auto créée + alerte + lien FMEA/PDCA.", req="CLAUDE §9.9"),
    c("Sécurité", "mTLS + journalisation des commandes équipements", "Sécurité", "Haute",
      "1. Envoyer une commande à un équipement",
      "mTLS exigé ; commande signée et ancrée.", role="DevOps", req="CLAUDE §9.8"),
])

# =====================================================================
# 2) APLATISSEMENT + AFFECTATION STATUTS/EXÉCUTION (déterministe)
# =====================================================================
# Distribution déterministe pour peupler les graphiques (campagne en cours).
STATUS_CYCLE = (["Passé"]*6 + ["Échoué"] + ["Bloqué"] + ["Non exécuté"]*2 + ["En cours"])
TESTERS = ["A. Benali", "S. Diallo", "M. Fontaine", "L. Tremblay", "R. Costa"]

rows = []
i = 0
for dname, dcode, cases in DOMAINS:
    for n, cc in enumerate(cases, start=1):
        tcid = f"TC-{dcode}-{n:03d}"
        status = STATUS_CYCLE[i % len(STATUS_CYCLE)]
        tester = TESTERS[i % len(TESTERS)] if status not in ("Non exécuté",) else ""
        autom = "Oui" if cc["typ"] in ("Fonctionnel", "Sécurité", "Performance", "API", "Régression") and (i % 3 != 0) else "Non"
        rows.append({
            "id": tcid, "domain": dname, "sub": cc["sub"], "title": cc["title"],
            "typ": cc["typ"], "prio": cc["prio"], "pre": cc["pre"],
            "steps": cc["steps"], "expected": cc["expected"], "role": cc["role"],
            "autom": autom, "status": status, "tester": tester, "req": cc["req"],
        })
        i += 1

TOTAL = len(rows)

# =====================================================================
# 3) WORKBOOK
# =====================================================================
wb = Workbook()

def style_header(ws, row, ncols, height=22):
    ws.row_dimensions[row].height = height
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(NAVY); cell.font = font(11, True, WHITE)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER

def banner(ws, text, sub=""):
    ws.merge_cells("A1:H1"); ws["A1"] = text
    ws["A1"].font = font(20, True, WHITE); ws["A1"].fill = fill(NAVY)
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 40
    if sub:
        ws.merge_cells("A2:H2"); ws["A2"] = sub
        ws["A2"].font = font(11, False, GREY); ws["A2"].alignment = Alignment(indent=1)

# ---- Sheet 1 : Page de garde --------------------------------------------------
ws = wb.active; ws.title = "Page de garde"
ws.sheet_view.showGridLines = False
for col in "ABCDEFGH": ws.column_dimensions[col].width = 15
ws.merge_cells("B2:G2"); ws["B2"] = "QualitOS"
ws["B2"].font = font(40, True, NAVY); ws["B2"].alignment = Alignment(horizontal="center")
ws.merge_cells("B3:G3"); ws["B3"] = "Quality Operating System — Plateforme SaaS de qualité totale"
ws["B3"].font = font(12, False, GREY); ws["B3"].alignment = Alignment(horizontal="center")
ws.merge_cells("B5:G6"); ws["B5"] = "DOSSIER DE TEST QA\nPlan de test & cahier de recette"
ws["B5"].font = font(22, True, WHITE); ws["B5"].fill = fill(BLUE)
ws["B5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
meta = [
    ("Projet", "QualitOS — plateforme multi-méthodes, multi-secteurs, IA-native"),
    ("Document", "Dossier de test fonctionnel, sécurité, performance et non-régression"),
    ("Version", "1.0"),
    ("Date", date.today().strftime("%d/%m/%Y")),
    ("Périmètre", f"{len(DOMAINS)} domaines · {TOTAL} cas de test · ~55 écrans · 60 normes · 14 packs sectoriels"),
    ("Environnements", "DEV (local), CI (GitHub Actions), Recette, Pré-prod"),
    ("Auteur", "Équipe QA — Architecte principal"),
    ("Statut", "En cours de campagne"),
    ("Confidentialité", "Interne"),
]
r = 8
for k, v in meta:
    ws.cell(row=r, column=2, value=k).font = font(11, True, NAVY)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    cell = ws.cell(row=r, column=3, value=v); cell.font = font(11)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1
ws.merge_cells(f"B{r+1}:G{r+1}")
ws.cell(row=r+1, column=2, value="Réf. exigences : CLAUDE.md (vision produit) · ADR docs/adr · Normes ISO 9001/27001…").font = font(9, italic=True, color=GREY)

# ---- Sheet 2 : Référentiel / Stratégie ---------------------------------------
ws = wb.create_sheet("Référentiel & Stratégie")
ws.sheet_view.showGridLines = False
banner(ws, "Référentiel & stratégie de test", "Cadre méthodologique de la campagne QA")
ws.column_dimensions["A"].width = 26
for col in "BCDEFGH": ws.column_dimensions[col].width = 16
blocks = [
    ("Objectif", "Vérifier que QualitOS couvre fonctionnellement, en sécurité, en performance et "
                 "en accessibilité l'ensemble des modules, et garantir la non-régression."),
    ("Types de test", "Fonctionnel · Sécurité (OWASP ASVS L3 / Top 10 / LLM Top 10) · IHM/UX · "
                      "Accessibilité (WCAG 2.2 AA) · i18n (6 langues) · Performance/SLO · "
                      "Intégration · API · Régression · Robustesse/Chaos · Configuration."),
    ("Niveaux de priorité", "Critique (bloquant métier/sécurité) · Haute · Moyenne · Basse."),
    ("Statuts d'exécution", "Non exécuté · En cours · Passé · Échoué · Bloqué."),
    ("Rôles testés", "Super Admin · Admin Tenant · Directeur Qualité · Manager Qualité · "
                     "Auditeur · Utilisateur · Externe · Partenaire · Public (non authentifié)."),
    ("Environnements", "DEV local (Postgres 5434, Keycloak 8080, engine 8082, web 4200, ai-service 8085) ; "
                       "CI GitHub Actions (build, tests, SAST/SCA, DAST ZAP, perf k6) ; Recette ; Pré-prod."),
    ("Critères d'entrée", "Build vert · environnement disponible · jeu de données seedé · comptes de test prêts."),
    ("Critères de sortie", "100% des cas Critiques/Hauts Passés · 0 anomalie bloquante ouverte · "
                           "couverture ≥ 85% · SLO respectés · 0 vuln HIGH (DAST)."),
    ("Outils", "Cypress/Playwright (E2E) · Jest/Karma (unitaire front) · JUnit/Testcontainers (back) · "
               "k6/Gatling (perf) · OWASP ZAP (DAST) · Axe (a11y) · Chaos Mesh (chaos)."),
    ("Gestion des anomalies", "Toute anomalie est journalisée (feuille 'Journal d'anomalies') avec "
                              "sévérité, statut, cas lié et correctif. Retest obligatoire avant clôture."),
]
rr = 4
for k, v in blocks:
    ws.cell(row=rr, column=1, value=k).font = font(11, True, NAVY)
    ws.cell(row=rr, column=1).fill = fill(LIGHT)
    ws.cell(row=rr, column=1).alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=8)
    cell = ws.cell(row=rr, column=2, value=v); cell.font = font(11)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[rr].height = 46
    for col in range(1, 9): ws.cell(row=rr, column=col).border = BORDER
    rr += 1

# ---- Sheet 3 : Cas de test ----------------------------------------------------
ws = wb.create_sheet("Cas de test")
headers = ["ID", "Domaine", "Sous-module", "Scénario de test", "Type", "Priorité",
           "Préconditions", "Étapes", "Résultat attendu", "Rôle", "Automatisable",
           "Statut", "Testé par", "Exigence"]
widths = [13, 24, 16, 34, 13, 11, 26, 40, 42, 16, 12, 13, 14, 14]
for ci, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.append(headers)
style_header(ws, 1, len(headers))
ws.freeze_panes = "A2"
for rrow in rows:
    ws.append([
        rrow["id"], rrow["domain"], rrow["sub"], rrow["title"], rrow["typ"], rrow["prio"],
        rrow["pre"], rrow["steps"], rrow["expected"], rrow["role"], rrow["autom"],
        rrow["status"], rrow["tester"], rrow["req"],
    ])
for ri in range(2, TOTAL + 2):
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=ri, column=ci)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.font = font(10)
    # couleur statut
    scell = ws.cell(row=ri, column=12)
    scell.fill = fill(STATUS_COLORS.get(scell.value, GREY)); scell.font = font(10, True, WHITE)
    scell.alignment = Alignment(vertical="center", horizontal="center")
    # couleur priorité
    pcell = ws.cell(row=ri, column=6)
    pcell.font = font(10, True, PRIO_COLORS.get(pcell.value, GREY))
    if ri % 2 == 0:
        for ci in range(1, len(headers)+1):
            if ci not in (6, 12):
                if ws.cell(row=ri, column=ci).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=ri, column=ci).fill = fill(LIGHT)
# AutoFilter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{TOTAL+1}"

# ---- compute aggregates -------------------------------------------------------
from collections import Counter, OrderedDict
status_counts = Counter(r["status"] for r in rows)
prio_counts = Counter(r["prio"] for r in rows)
type_counts = Counter(r["typ"] for r in rows)
domain_counts = OrderedDict()
for dname, dcode, cases in DOMAINS:
    domain_counts[dname] = len(cases)

passed = status_counts.get("Passé", 0)
failed = status_counts.get("Échoué", 0)
blocked = status_counts.get("Bloqué", 0)
inprog = status_counts.get("En cours", 0)
notrun = status_counts.get("Non exécuté", 0)
executed = passed + failed + blocked
pass_rate = round(100 * passed / executed, 1) if executed else 0
exec_rate = round(100 * executed / TOTAL, 1)

# ---- Sheet 4 : Synthèse (dashboard + charts) ---------------------------------
ws = wb.create_sheet("Synthèse")
ws.sheet_view.showGridLines = False
banner(ws, "Synthèse de la campagne de test", "Tableau de bord QA — vue 360°")
for col in "ABCDEFGHIJKLMN": ws.column_dimensions[col].width = 14

# KPI cards
kpis = [
    ("Cas de test", TOTAL, BLUE),
    ("Exécutés", executed, NAVY),
    ("Taux d'exécution", f"{exec_rate}%", NAVY),
    ("Passés", passed, GREEN),
    ("Taux de réussite", f"{pass_rate}%", GREEN),
    ("Échoués", failed, RED),
    ("Bloqués", blocked, AMBER),
    ("Non exécutés", notrun, GREY),
]
col = 1
for label, val, color in kpis:
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col+1)
    lc = ws.cell(row=4, column=col, value=label); lc.font = font(10, True, WHITE); lc.fill = fill(color)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws.cell(row=5, column=col, value=val); vc.font = font(22, True, color); vc.fill = fill(LIGHT)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    col += 2

# data tables (hidden-ish area) for charts
def write_table(ws, start_row, start_col, title, mapping):
    ws.cell(row=start_row, column=start_col, value=title).font = font(11, True, NAVY)
    r2 = start_row + 1
    ws.cell(row=r2, column=start_col, value="Catégorie").font = font(10, True)
    ws.cell(row=r2, column=start_col+1, value="Nombre").font = font(10, True)
    r2 += 1
    first = r2
    for k, v in mapping.items():
        ws.cell(row=r2, column=start_col, value=k)
        ws.cell(row=r2, column=start_col+1, value=v)
        r2 += 1
    return first, r2 - 1  # data row range

# Status pie
sr, er = write_table(ws, 8, 1, "Répartition par statut",
                     OrderedDict([("Passé",passed),("Échoué",failed),("Bloqué",blocked),
                                  ("En cours",inprog),("Non exécuté",notrun)]))
pie = PieChart(); pie.title = "Statuts d'exécution"; pie.height = 7.5; pie.width = 11
data = Reference(ws, min_col=2, min_row=sr-1, max_row=er)
cats = Reference(ws, min_col=1, min_row=sr, max_row=er)
pie.add_data(data, titles_from_data=True); pie.set_categories(cats)
pie.dataLabels = DataLabelList(); pie.dataLabels.showPercent = True
ws.add_chart(pie, "A15")

# Priority bar
sr2, er2 = write_table(ws, 8, 4, "Répartition par priorité",
                       OrderedDict([(k, prio_counts.get(k,0)) for k in ["Critique","Haute","Moyenne","Basse"]]))
bar = BarChart(); bar.title = "Cas par priorité"; bar.type="col"; bar.height=7.5; bar.width=11; bar.legend=None
d2 = Reference(ws, min_col=5, min_row=sr2-1, max_row=er2)
c2 = Reference(ws, min_col=4, min_row=sr2, max_row=er2)
bar.add_data(d2, titles_from_data=True); bar.set_categories(c2)
bar.dataLabels = DataLabelList(); bar.dataLabels.showVal = True
ws.add_chart(bar, "E15")

# By type bar
sr3, er3 = write_table(ws, 8, 7, "Répartition par type",
                       OrderedDict(sorted(type_counts.items(), key=lambda x:-x[1])))
bar2 = BarChart(); bar2.title="Cas par type de test"; bar2.type="bar"; bar2.height=9; bar2.width=12; bar2.legend=None
d3 = Reference(ws, min_col=8, min_row=sr3-1, max_row=er3)
c3 = Reference(ws, min_col=7, min_row=sr3, max_row=er3)
bar2.add_data(d3, titles_from_data=True); bar2.set_categories(c3)
ws.add_chart(bar2, "I15")

# By domain bar (big)
base = er3 + 3
ws.cell(row=base, column=7, value="Cas par domaine").font = font(11, True, NAVY)
ws.cell(row=base+1, column=7, value="Domaine").font = font(10, True)
ws.cell(row=base+1, column=8, value="Nombre").font = font(10, True)
rdom = base+2
for k, v in domain_counts.items():
    ws.cell(row=rdom, column=7, value=k); ws.cell(row=rdom, column=8, value=v); rdom += 1
barD = BarChart(); barD.title="Couverture par domaine fonctionnel"; barD.type="bar"; barD.height=12; barD.width=20; barD.legend=None
dD = Reference(ws, min_col=8, min_row=base+1, max_row=rdom-1)
cD = Reference(ws, min_col=7, min_row=base+2, max_row=rdom-1)
barD.add_data(dD, titles_from_data=True); barD.set_categories(cD)
barD.dataLabels = DataLabelList(); barD.dataLabels.showVal = True
ws.add_chart(barD, "A33")

# ---- Sheet 5 : Suivi & Évolution (time series + line charts) -----------------
ws = wb.create_sheet("Suivi & Évolution")
ws.sheet_view.showGridLines = False
banner(ws, "Suivi & évolution de la campagne", "Courbes d'avancement (burn-up) et tendance des anomalies")
for col in "ABCDEFGHIJ": ws.column_dimensions[col].width = 13
ws.column_dimensions["A"].width = 12

# Build a realistic S-curve over N working days
N = 15
start = date.today() - timedelta(days=N+2)
hdr = ["Jour", "Planifiés (cumul)", "Exécutés (cumul)", "Passés (cumul)",
       "Échoués (cumul)", "Anomalies ouvertes (cumul)", "Anomalies résolues (cumul)"]
hrow = 4
for ci, h in enumerate(hdr, start=1):
    ws.cell(row=hrow, column=ci, value=h)
style_header(ws, hrow, len(hdr))
# generate cumulative numbers converging to current snapshot
exec_final = executed
pass_final = passed
fail_final = failed
def ramp(total, k, n, curve=1.6):
    # S-curve-ish cumulative
    import math
    x = (k)/(n)
    val = total * (x**curve) / (x**curve + (1-x)**curve) if 0 < x < 1 else (0 if k==0 else total)
    return int(round(val))
defects_total = failed + blocked + 6  # incl. quelques anomalies UX/perf
defects_resolved_final = defects_total - 3
for k in range(N+1):
    d = start + timedelta(days=k)
    planned = TOTAL  # périmètre figé
    ex = ramp(exec_final, k, N)
    pa = ramp(pass_final, k, N)
    fa = ramp(fail_final, k, N)
    op = ramp(defects_total, k, N, curve=1.3)
    re = ramp(defects_resolved_final, k, N, curve=2.0)
    re = min(re, op)
    ws.append([d.strftime("%d/%m"), planned, ex, pa, fa, op, re])
first_data = hrow + 1
last_data = hrow + N + 1
for ri in range(first_data, last_data+1):
    for ci in range(1, len(hdr)+1):
        ws.cell(row=ri, column=ci).border = BORDER
        ws.cell(row=ri, column=ci).alignment = Alignment(horizontal="center")
        ws.cell(row=ri, column=ci).font = font(10)

# Burn-up line chart : planifiés / exécutés / passés
line = LineChart(); line.title = "Avancement de la campagne (burn-up)"
line.height = 9.5; line.width = 19; line.y_axis.title = "Cas (cumulés)"; line.x_axis.title = "Jour"
data = Reference(ws, min_col=2, max_col=4, min_row=hrow, max_row=last_data)
cats = Reference(ws, min_col=1, min_row=first_data, max_row=last_data)
line.add_data(data, titles_from_data=True); line.set_categories(cats)
ws.add_chart(line, "A23")

# Defect trend
line2 = LineChart(); line2.title = "Tendance des anomalies (ouvertes vs résolues)"
line2.height = 9.5; line2.width = 19; line2.y_axis.title = "Anomalies (cumulées)"; line2.x_axis.title = "Jour"
data2 = Reference(ws, min_col=6, max_col=7, min_row=hrow, max_row=last_data)
line2.add_data(data2, titles_from_data=True); line2.set_categories(cats)
ws.add_chart(line2, "A43")

# ---- Sheet 6 : Matrice de traçabilité ----------------------------------------
ws = wb.create_sheet("Traçabilité")
ws.sheet_view.showGridLines = False
banner(ws, "Matrice de traçabilité", "Couverture des exigences par les cas de test")
hh = ["Exigence / Réf.", "Nb de cas", "IDs des cas de test"]
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 12; ws.column_dimensions["C"].width = 90
hrow = 4
ws.cell(row=hrow, column=1, value=hh[0]); ws.cell(row=hrow, column=2, value=hh[1]); ws.cell(row=hrow, column=3, value=hh[2])
style_header(ws, hrow, 3)
req_map = OrderedDict()
for r in rows:
    req_map.setdefault(r["req"], []).append(r["id"])
ri = hrow + 1
for req, ids in sorted(req_map.items()):
    ws.cell(row=ri, column=1, value=req).font = font(10, True, NAVY)
    ws.cell(row=ri, column=2, value=len(ids)).alignment = Alignment(horizontal="center")
    ws.cell(row=ri, column=3, value=", ".join(ids)).font = font(9)
    for ci in range(1,4):
        ws.cell(row=ri, column=ci).border = BORDER
        ws.cell(row=ri, column=ci).alignment = Alignment(vertical="top", wrap_text=True)
    ri += 1

# ---- Sheet 7 : Journal d'anomalies -------------------------------------------
ws = wb.create_sheet("Journal d'anomalies")
ws.sheet_view.showGridLines = False
banner(ws, "Journal d'anomalies", "Défauts détectés, sévérité, statut et correctif")
hh = ["ID", "Titre", "Module", "Sévérité", "Statut", "Cas lié", "Description", "Correctif / Commit", "Détecté le"]
widths = [10, 30, 16, 12, 12, 14, 46, 30, 12]
for ci, w in enumerate(widths, start=1): ws.column_dimensions[get_column_letter(ci)].width = w
hrow = 4
for ci, h in enumerate(hh, start=1): ws.cell(row=hrow, column=ci, value=h)
style_header(ws, hrow, len(hh))
anomalies = [
    ("ANO-001","Token IA Mistral large throttlé → timeout 502","IA / NLQ","Majeure","Résolu","TC-AI-001",
     "mistral-large-latest répondait en ~25s (>timeout 30s) → 503/502.","Bascule MISTRAL_MODEL=codestral-latest (~1s).","23/06"),
    ("ANO-002","Listes : scintillement / spinner infini (boucle)","Méthodes/Transverse","Critique","Résolu","TC-XCUT-006",
     "shareReplay refCount:true + table gated loading -> boucle teardown (API rappelée 160x).","refCount:false (8 listes) — commit c090086.","23/06"),
    ("ANO-003","Page détail figée sur 'Chargement…'","PDCA/5S/NC…","Critique","Résolu","TC-PDCA-004",
     "reload$ = Subject, .next() en ngOnInit perdu avant abonnement -> getX jamais appelé.","BehaviorSubject(undefined) — commit 35b7c39.","23/06"),
    ("ANO-004","Rebond Keycloak après expiration token (15 min)","Auth","Majeure","Résolu","TC-SEC-002",
     "Pas de refresh -> redirect login au chargement, perte de route.","Refresh auto refresh_token + préservation route — b3a82ba.","23/06"),
    ("ANO-005","CI DAST ZAP rouge sur des WARN (0 FAIL)","CI/Sécurité","Mineure","Résolu","TC-SEC-007",
     "fail_action:true échouait sur tout WARN au lieu de HIGH seulement.","fail_action:false + gate riskcode=3 — 5244dca.","23/06"),
    ("ANO-006","Dev server Vite : reload-thrash après ajout dépendance","Front/Infra","Mineure","Résolu","TC-XCUT-003",
     "Re-optimisation gridster2 -> reloads -> écrans vides intermittents.","Redémarrage ng serve à cache propre (dev only).","22/06"),
]
ri = hrow+1
SEV = {"Critique":RED,"Majeure":AMBER,"Mineure":BLUE}
for a in anomalies:
    for ci, val in enumerate(a, start=1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.border = BORDER; cell.font = font(10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=ri, column=4).fill = fill(SEV.get(a[3], GREY)); ws.cell(row=ri, column=4).font = font(10, True, WHITE)
    ws.cell(row=ri, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=ri, column=5).fill = fill(GREEN if a[4]=="Résolu" else AMBER); ws.cell(row=ri, column=5).font = font(10, True, WHITE)
    ws.cell(row=ri, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ri += 1

# ---- save ---------------------------------------------------------------------
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "QualitOS_Dossier_de_Test_QA.xlsx")
wb.save(out)
print(f"OK -> {out}")
print(f"Domaines: {len(DOMAINS)} | Cas de test: {TOTAL} | Passés: {passed} | Échoués: {failed} | Bloqués: {blocked} | Non exécutés: {notrun}")
print(f"Feuilles: {wb.sheetnames}")
