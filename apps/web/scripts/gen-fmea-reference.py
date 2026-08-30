# -*- coding: utf-8 -*-
"""Genere src/app/features/fmea/fmea.reference.ts depuis le classeur de reference.

Usage : python scripts/gen-fmea-reference.py            (depuis apps/web)
        python scripts/gen-fmea-reference.py --check    (verifie sans ecrire)

Source : docs/QUALITOS BACKLOG.xlsx
  * feuille 4 -> les trois baremes (Severite / Detection / Occurrence)
  * feuille 2 -> l'exemple complet de PFMEA fourni comme modele de redaction

Pourquoi un generateur plutot qu'une recopie : trente lignes chiffrees recopiees
a la main, c'est une erreur de transcription qui passe inapercue et qui fausse
des cotations. Le classeur reste la source ; ce fichier n'en est que la sortie.

Dependance : openpyxl (lecture seule du classeur).
"""
import io
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, '..', '..', '..'))
WORKBOOK = os.path.join(REPO, 'docs', 'QUALITOS BACKLOG.xlsx')
TARGET = os.path.join(HERE, '..', 'src', 'app', 'features', 'fmea', 'fmea.reference.ts')

SCALES_SHEET = 'Feuille 4'
EXAMPLE_SHEET = 'Feuille 2'

# Les baremes occupent trois blocs de colonnes cote a cote, lignes 5 a 14.
SCALE_ROWS = range(5, 15)
# L'exemple de PFMEA commence sous ses deux lignes d'en-tete.
EXAMPLE_ROWS = range(12, 27)
EXAMPLE_COLS = range(2, 13)

HEADER = '''/**
 * Referentiel FMEA - extrait du classeur de reference qualite (baremes de
 * cotation, feuille 4, et exemple de PFMEA, feuille 2).
 *
 * <p>Ces tables ne sont pas des donnees de tenant : ce sont les echelles sur
 * lesquelles se cotent Severite, Occurrence et Detection. Sans elles a portee
 * d'ecran, chaque evaluateur invente son propre 1 a 10 et les RPN de deux
 * analyses cessent d'etre comparables.
 *
 * <p>Le contenu est reproduit dans la langue du referentiel d'origine
 * (anglais) : une echelle de cotation traduite librement n'est plus la meme
 * echelle. Seule l'interface qui l'entoure est traduite.
 *
 * <p>FICHIER GENERE par scripts/gen-fmea-reference.py depuis
 * « docs/QUALITOS BACKLOG.xlsx » - ne pas editer a la main : corriger le
 * classeur, puis regenerer.
 */

/** Une ligne du bareme de severite. */
export interface FmeaSeverityRow { effect: string; description: string; score: number; }

/** Une ligne du bareme de detection. */
export interface FmeaDetectionRow { chance: string; description: string; score: number; }

/**
 * Une ligne du bareme d'occurrence. `probability` est vide sur les lignes que
 * le referentiel regroupe sous l'intitule precedent - on ne complete pas ce
 * qu'il laisse en blanc.
 */
export interface FmeaOccurrenceRow {
  probability: string; timePeriod: string; failureRate: string; score: number;
}

/** Une ligne de l'exemple de PFMEA fourni comme modele. */
export interface FmeaExampleRow {
  step: string; failureMode: string; effects: string; severity: number;
  causes: string; occurrence: number; controls: string; detection: number;
  rpn: number; recommendedAction: string; responsible: string;
}
'''

EXAMPLE_TITLE = 'Electrical Wiring Harness - Aircraft Installation (PFMEA)'


def text(value):
    """Chaine TS, espaces normalises. Une cellule vide donne une chaine vide."""
    if value is None:
        return "''"
    return json.dumps(' '.join(str(value).split()), ensure_ascii=False)


def score(value):
    """Les cotations sont saisies en flottants dans le classeur ; ce sont des entiers."""
    return str(int(value)) if value is not None else '0'


def render():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    scales = workbook[SCALES_SHEET]
    example = workbook[EXAMPLE_SHEET]

    out = [HEADER]

    out.append("/** Severite : gravite de l'effet pour le client, de 10 (danger) a 1 (aucun effet). */")
    out.append('export const FMEA_SEVERITY_SCALE: ReadonlyArray<FmeaSeverityRow> = [')
    for row in SCALE_ROWS:
        out.append('  { effect: %s, description: %s, score: %s },' % (
            text(scales.cell(row, 1).value),
            text(scales.cell(row, 2).value),
            score(scales.cell(row, 3).value)))
    out.append('];\n')

    out.append('/** Detection : chance de reperer la defaillance avec les controles en place. */')
    out.append('export const FMEA_DETECTION_SCALE: ReadonlyArray<FmeaDetectionRow> = [')
    for row in SCALE_ROWS:
        out.append('  { chance: %s, description: %s, score: %s },' % (
            text(scales.cell(row, 5).value),
            text(scales.cell(row, 6).value),
            score(scales.cell(row, 7).value)))
    out.append('];\n')

    out.append('/** Occurrence : frequence attendue de la defaillance. */')
    out.append('export const FMEA_OCCURRENCE_SCALE: ReadonlyArray<FmeaOccurrenceRow> = [')
    for row in SCALE_ROWS:
        out.append('  { probability: %s, timePeriod: %s, failureRate: %s, score: %s },' % (
            text(scales.cell(row, 9).value),
            text(scales.cell(row, 10).value),
            text(scales.cell(row, 11).value),
            score(scales.cell(row, 12).value)))
    out.append('];\n')

    out.append('''/**
 * Exemple complet de PFMEA (faisceau electrique aeronautique), fourni comme
 * modele de redaction : ce qu'on attend dans « mode de defaillance », dans
 * « effets », et a quoi ressemble une action recommandee qui engage vraiment.
 */''')
    out.append('export const FMEA_EXAMPLE_TITLE = %s;' % text(EXAMPLE_TITLE))
    out.append('export const FMEA_EXAMPLE_ROWS: ReadonlyArray<FmeaExampleRow> = [')
    fields = ['step', 'failureMode', 'effects', 'severity', 'causes', 'occurrence',
              'controls', 'detection', 'rpn', 'recommendedAction', 'responsible']
    numeric = {'severity', 'occurrence', 'detection', 'rpn'}
    for row in EXAMPLE_ROWS:
        values = [example.cell(row=row, column=c).value for c in EXAMPLE_COLS]
        if values[0] is None:
            continue                      # ligne vide de fin de tableau
        out.append('  {')
        for i, field in enumerate(fields):
            rendered = score(values[i]) if field in numeric else text(values[i])
            comma = '' if i == len(fields) - 1 else ','
            out.append('    %s: %s%s' % (field, rendered, comma))
        out.append('  },')
    out.append('];')

    return '\n'.join(out) + '\n'


def main():
    generated = render()
    check = '--check' in sys.argv

    if check:
        current = io.open(TARGET, encoding='utf-8').read()
        if current != generated:
            print('DESYNCHRONISE : %s ne correspond plus au classeur.' % TARGET)
            print('Regenerer avec : python scripts/gen-fmea-reference.py')
            return 1
        print('OK %s est a jour' % TARGET)
        return 0

    io.open(TARGET, 'w', encoding='utf-8', newline='\n').write(generated)
    print('OK %s regenere' % TARGET)
    return 0


if __name__ == '__main__':
    sys.exit(main())
